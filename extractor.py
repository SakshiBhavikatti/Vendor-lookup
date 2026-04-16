from openpyxl import load_workbook
from preprocess import normalize_text


def load_vendors(file_path="data/vendors.xlsx"):
    wb = load_workbook(file_path, data_only=True)
    sheet = wb.active

    headers = [cell.value for cell in sheet[1]]
    data = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = dict(zip(headers, row))

        if not any(row_dict.values()):
            continue

        vendor_code = str(row_dict.get("vendor_code", "")).strip()
        vendor_name = normalize_text(row_dict.get("vendor_name", ""))

        if not vendor_code or not vendor_name:
            continue

        data.append({
            "vendor_code": vendor_code,
            "vendor_name": vendor_name
        })

    return data