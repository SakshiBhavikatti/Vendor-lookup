from openpyxl import load_workbook
from preprocess import normalize_text


def load_line_items(file_path="data/vendor_invoice_register (1).xlsx"):
    wb = load_workbook(file_path, data_only=True)
    sheet = wb["Vendor Invoice Register"]

    # Row 1 = group labels (HEADER INFORMATION / LINE ITEM INFORMATION) — skip
    # Row 2 = actual column headers
    headers = [
        str(cell.value).strip().lower().replace(" ", "_").replace("(₹)", "").strip("_")
        if cell.value else ""
        for cell in sheet[2]
    ]

    data = []

    for row in sheet.iter_rows(min_row=3, values_only=True):
        row_dict = dict(zip(headers, row))

        if not any(row_dict.values()):
            continue

        line_item = str(row_dict.get("line_item", "")).strip()
        if not line_item:
            continue

        data.append({
            # --- HEADER INFORMATION ---
            "vendor_code":       str(row_dict.get("vendor_code", "")).strip(),
            "vendor_name":       normalize_text(row_dict.get("vendor_name", "")),
            "company_code":      str(row_dict.get("company_code", "")).strip(),
            "business_place":    str(row_dict.get("business_place", "")).strip(),
            "section_code":      str(row_dict.get("section_code", "")).strip(),
            "total_invoice_amt": float(row_dict.get("total_invoice_amt") or 0),
            # --- LINE ITEM INFORMATION ---
            "line_item":                 line_item,
            "material_code":             str(row_dict.get("material_code", "")).strip(),
            "description":               normalize_text(row_dict.get("description", "")),
            "service_code":              str(row_dict.get("service_code", "")).strip(),
            "ocr_line_item_description": normalize_text(row_dict.get("ocr_line_item_description", "")),
            "line_amt":                  float(row_dict.get("line_amt") or 0),
            "cost_centre":               str(row_dict.get("cost_centre", "")).strip(),
            "gl_code":                   str(row_dict.get("gl_code", "")).strip(),
        })

    return data